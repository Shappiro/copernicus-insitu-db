import datetime
import json
import string
from io import BytesIO

from django.template.loader import get_template
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404
from django.views import View

from explorer.app_settings import (
    ENABLE_TASKS,
    EXPLORER_DEFAULT_ROWS,
    UNSAFE_RENDERING,
)
from explorer.exporters import get_exporter_class
from explorer.forms import QueryForm
from explorer.models import Query
from explorer.views import (
    DownloadQueryView,
    PlayQueryView,
    _export,
    query_viewmodel
)
from explorer.utils import extract_params
from explorer.utils import url_get_rows

from wkhtmltopdf.views import PDFTemplateResponse, PDFTemplateView

from insitu.views import protected
from insitu.views.protected.views import ProtectedTemplateView, ProtectedView


def as_text(value):
    if value is None:
        return ""
    return str(value)


class ReportsListView(ProtectedTemplateView):
    template_name = 'reports/list.html'
    permission_classes = (protected.IsAuthenticated,)
    permission_denied_redirect = reverse_lazy('auth:login')

    def get_context_data(self, **kwargs):
        context = super(ReportsListView, self).get_context_data(**kwargs)
        context['queries'] = Query.objects.all().order_by('id').values(
            'id', 'title', 'description')
        return context


class ReportsDetailView(ProtectedTemplateView):
    template_name = 'reports/detail.html'
    permission_classes = (protected.IsAuthenticated,)
    permission_denied_redirect = reverse_lazy('auth:login')

    def get(self, request, *args, **kwargs):
        self.report = get_object_or_404(Query, pk=kwargs['query_id'])
        return super(ReportsDetailView, self).get(request, *args, **kwargs)

    def get_filename(self, title):
        valid_chars = '-_.() %s%s' % (string.ascii_letters, string.digits)
        filename = ''.join(c for c in title if c in valid_chars)
        filename = filename.replace(' ', '_')
        return filename

    def get_context_data(self, **kwargs):
        context = super(ReportsDetailView, self).get_context_data(**kwargs)
        res = self.report.execute_query_only()
        context['query'] = {
            'id': self.report.id,
            'title': self.report.title,
            'description': self.report.description,
            'params': extract_params(self.report.sql),
        }
        filename =  (
            self.get_filename(self.report.title) +
            datetime.datetime.now().strftime('%Y%m%d')
        )
        context.update({
            'html_filename': filename + '.html',
            'pdf_filename': filename + '.pdf',
            'excel_filename': filename + '.xlsx',
            'no_jquery': True,
            'unsafe_rendering': UNSAFE_RENDERING,
        })
        return context

class ReportDataJsonView(ProtectedView):

    def get(self, request, *args, **kwargs):
        self.report = get_object_or_404(Query, pk=kwargs['query_id'])
        res = self.report.execute_query_only()
        data = []
        for row in res.data:
            column_count = 0
            row_data = {}
            for column in res.headers:
                row_data[column.title] = str(row[column_count])
                column_count = column_count + 1
            data.append(row_data)

        return JsonResponse(data, safe=False)

class PlaygroundView(PlayQueryView):

    def render(self):
        return self.render_template(
            'reports/playground.html',
            {'title': 'Playground', 'form': QueryForm(),
             'no_jquery': True}
        )

    def render_with_sql(self, request, query, run_query=True, error=None):
        rows = url_get_rows(request)
        context = query_viewmodel(request.user, query,
                                  title="Playground",
                                  run_query=run_query,
                                  error=error, rows=rows)
        context.update({'no_jquery': True})
        return self.render_template('reports/playground.html', context)


class SnapshotView(ProtectedTemplateView):

    def get(self, request, *args, **kwargs):
        response = HttpResponse()
        date = datetime.datetime.now().strftime('%Y%m%d')
        response["Content-Disposition"] = "attachment; filename=insitu_{0}.sql.gz".format(
            date)
        response['X-Accel-Redirect'] = "/static/protected/database.sql.gz"
        return response


class DownloadReportsView(DownloadQueryView):

    def get(self, request, query_id, *args, **kwargs):
        query = get_object_or_404(Query, pk=query_id)
        format = request.GET.get('format', 'csv')
        exporter_class = get_exporter_class(format)
        file_extension = exporter_class.file_extension
        date = '_' + datetime.datetime.now().strftime('%Y%m%d')
        response = _export(request, query)
        response['Content-Disposition'] = (date + file_extension).join(
            response['Content-Disposition'].split(file_extension))
        if file_extension == '.xlsx':
            wb = load_workbook(filename=BytesIO(response.content))
            ws = wb.active
            dims = {}
            for row in ws.iter_rows():
                for cell in row:
                    dims[cell.column] = max(dims.get(cell.column, 0),
                                            len(as_text(cell.value)))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            wb.close()
            virtual_wb = save_virtual_workbook(wb)
            response.content = virtual_wb
        return response


class Pdf(View):

    def render(self, context, request):
        template = get_template('reports/reports_pdf.html')
        template_response = PDFTemplateResponse(
            request=request,
            template=template,
            filename='test.pdf',
            context=context,
            show_content_in_browser=False,
            cmd_options={
                'javascript-delay': 3000,
            },
        )
        template_response.render()
        content = template_response.rendered_content
        return HttpResponse(content, content_type='application/pdf')

    def post(self, request, *args, **kwargs):
        context = {
            "data": request.POST['data'],
            "title": request.POST['title'],
            "date": datetime.datetime.now().strftime('%Y %m %d'),
        }
        return self.render(context, request)
